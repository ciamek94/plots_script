import os
import pandas as pd
import folium
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import requests
import json
from datetime import date, datetime
from dotenv import load_dotenv

# Import skryptów źródłowych
from otodom import main as main_script1
from olx import main as main_script2
from nieruchomosci_online import main as main_script3

# -------------------------------
# 🔧 Load secrets from .env
load_dotenv()

# -------------------------------
# 🔐 OneDrive auth
CLIENT_ID = os.environ['ONEDRIVE_CLIENT_ID']
REFRESH_TOKEN = os.environ['ONEDRIVE_REFRESH_TOKEN']
SCOPES = ['offline_access', 'Files.ReadWrite.All']
TOKEN_URL = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'

# -------------------------------
# 📂 Pliki lokalne
EXCEL_FILE_1 = os.path.join('dzialki', 'otodom_dzialki.xlsx')
EXCEL_FILE_2 = os.path.join('dzialki', 'olx_dzialki.xlsx')
EXCEL_FILE_3 = os.path.join('dzialki', 'nieruchomosci_online_dzialki.xlsx')

EXCEL_MERGED = os.path.join('dzialki_merged', 'dzialki_merged.xlsx')
MAP_MERGED = os.path.join('dzialki_merged', 'map_merged.html')
SENT_JSON = os.path.join('dzialki_merged', 'sent_ads.json')

# -------------------------------
# 🤖 Telegram
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")

# -------------------------------
# 🔐 OneDrive token refresh
def authenticate():
    """Authenticate with OneDrive API using refresh token"""
    data = {
        'client_id': CLIENT_ID,
        'refresh_token': REFRESH_TOKEN,
        'grant_type': 'refresh_token',
        'scope': 'offline_access Files.ReadWrite.All',
    }
    resp = requests.post(TOKEN_URL, data=data)
    if resp.status_code != 200:
        raise Exception(f"❌ Failed to authenticate: {resp.text}")
    return resp.json()

# -------------------------------
# ☁️ Upload file to OneDrive
def upload_to_onedrive(file_path, token):
    """Upload a file to OneDrive root directory"""
    headers = {
        'Authorization': f"Bearer {token['access_token']}",
        'Content-Type': 'application/octet-stream'
    }
    with open(file_path, 'rb') as f:
        file_data = f.read()
    upload_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{file_path}:/content'
    r = requests.put(upload_url, headers=headers, data=file_data)
    if r.status_code in (200, 201):
        print(f"✅ File uploaded to OneDrive: {file_path}")
    else:
        print(f"❌ Upload failed: {r.status_code} {r.text}")

# -------------------------------
# ☁️ Download file from OneDrive
def download_from_onedrive(file_path, token):
    """Download a file from OneDrive if it exists"""
    headers = {'Authorization': f"Bearer {token['access_token']}"}
    url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{file_path}:/content'
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        with open(file_path, "wb") as f:
            f.write(r.content)
        print(f"⬇️ Downloaded from OneDrive: {file_path}")
        return True
    else:
        print(f"⚠️ No remote file found for {file_path} (status {r.status_code})")
        return False

# -------------------------------
# 📦 Sent ads JSON
def load_sent_ads():
    """Load sent ads from JSON file"""
    if os.path.exists(SENT_JSON):
        with open(SENT_JSON, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()

def save_sent_ads(sent_ads):
    """Save sent ads to JSON file"""
    with open(SENT_JSON, "w", encoding="utf-8") as f:
        json.dump(list(sent_ads), f, ensure_ascii=False, indent=2)

# -------------------------------
# 📲 Telegram
def send_telegram_message(title, link, price, image_url=None):
    """Send a message with optional photo to Telegram"""
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("⚠️ Telegram not configured, skipping notification")
        return

    message = f"🏡 <b>{title}</b>\n💰 {price}\n🔗 <a href='{link}'>Zobacz ogłoszenie</a>"
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    data = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": message,
        "parse_mode": "HTML"
    }
    requests.post(url, data=data)

    if image_url:
        url_photo = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto"
        data_photo = {"chat_id": TELEGRAM_CHAT_ID, "photo": image_url, "caption": title}
        requests.post(url_photo, data=data_photo)

# -------------------------------
# 🔄 Merge Excels
def merge_excels(files_list, output_file):
    dfs = []
    for idx, file in enumerate(files_list):
        if not os.path.exists(file):
            print(f"❗ File not found: {file}")
            continue

        if idx == 0:
            xls = pd.ExcelFile(file)
            try:
                df_krakow = pd.read_excel(xls, sheet_name='powiat krakowski')
                df_wielicki = pd.read_excel(xls, sheet_name='powiat wielicki')
                df = pd.concat([df_krakow, df_wielicki], ignore_index=True)
            except Exception as e:
                print(f"Error reading sheets in {file}: {e}")
                df = pd.read_excel(file)
        else:
            df = pd.read_excel(file)

        source_name = ['otodom', 'olx', 'nieruchomosci-online'][idx]
        df['Source'] = source_name
        dfs.append(df)

    if not dfs:
        print("❗ Brak plików do scalania!")
        return None

    df_combined = pd.concat(dfs, ignore_index=True)
    df_unique = df_combined.reset_index(drop=True)

    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    df_unique.to_excel(output_file, index=False)

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    for col_cells in ws.columns:
        max_length = 0
        col = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col].width = max_length + 2
    wb.save(output_file)
    print(f"💾 Merged Excel saved: {output_file}")
    return df_unique

# -------------------------------
# 🗺️ Map generation
def generate_merged_map(df, map_path):
    KRAKOW_COORDS = (50.0647, 19.9450)
    m = folium.Map(location=KRAKOW_COORDS, zoom_start=10)

    folium.Marker(
        location=KRAKOW_COORDS,
        popup="<b>Kraków</b><br>Reference point",
        tooltip="Kraków",
        icon=folium.Icon(color="purple", icon="star", prefix="fa")
    ).add_to(m)

    location_to_listings = defaultdict(list)

    for _, row in df.iterrows():
        if 'Active' in df.columns and not row.get("Active", True):
            continue
        lat, lon = row.get("Latitude"), row.get("Longitude")
        if pd.isna(lat) or pd.isna(lon):
            continue
        coord_key = (round(lat, 5), round(lon, 5))
        location_to_listings[coord_key].append(row)

    for (lat, lon), listings in location_to_listings.items():
        if len(listings) == 1:
            l = listings[0]
            popup_html = f"""
                <b>{l.get('Title', 'Brak tytułu')}</b><br>
                {l.get('Location', '')}<br>
                {l.get('Price last updated', l.get('Price at first find', l.get('Price', '?')))}<br>
                <a href="{l.get('Link', '#')}" target="_blank">View listing</a>
            """
            tooltip = l.get("Title", "Listing")
            source = l.get("Source", "").lower()
        else:
            popup_html = f"<b>{len(listings)} ogłoszenia</b><br><ul>"
            for l in listings:
                price = l.get('Price last updated', l.get('Price at first find', l.get('Price', '?')))
                popup_html += f"<li><a href='{l.get('Link', '#')}' target='_blank'>{l.get('Title', 'Brak tytułu')}</a> – {price}</li>"
            popup_html += "</ul>"
            tooltip = f"{len(listings)} ogłoszeń"
            source = listings[0].get("Source", "").lower()

        if source == 'otodom':
            color = "green"
        elif source == 'olx':
            color = "blue"
        elif source == 'nieruchomosci-online':
            color = "orange"
        else:
            color = "gray"

        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_html, max_width=300),
            tooltip=tooltip,
            icon=folium.Icon(color=color, icon="home", prefix="fa")
        ).add_to(m)

    m.save(map_path)
    print(f"🗺️ Merged map saved: {map_path}")

# -------------------------------
# 🚀 Main
def main():
    token = authenticate()

    # ⬇️ Najpierw pobierz sent_ads.json z OneDrive (jeśli istnieje)
    download_from_onedrive(SENT_JSON, token)

    print("🚀 Uruchamiam skrypt 1 (Otodom)...")
    main_script1()

    print("🚀 Uruchamiam skrypt 2 (OLX)...")
    main_script2()

    print("🚀 Uruchamiam skrypt 3 (Nieruchomosci-online)...")
    main_script3()

    print("🔄 Scalanie wszystkich plików Excel...")
    df_merged = merge_excels([EXCEL_FILE_1, EXCEL_FILE_2, EXCEL_FILE_3], EXCEL_MERGED)
    if df_merged is not None:
        print("🗺️ Tworzenie mapy z danych scalonych...")
        generate_merged_map(df_merged, MAP_MERGED)

    print(f"📦 Upload {EXCEL_MERGED}, map and sent_ads.json to OneDrive...")
    upload_to_onedrive(EXCEL_MERGED, token)
    upload_to_onedrive(MAP_MERGED, token)

    today = date.today().strftime("%Y-%m-%d")
    sent_ads = load_sent_ads()
    current_ads = set()

    for _, row in df_merged.iterrows():
        link = row.get("Link")
        if not link:
            continue
        current_ads.add(link)

        # ✅ Normalizacja daty
        date_first = str(row.get("Date first found"))
        try:
            if "." in date_first:
                parsed_date = datetime.strptime(date_first, "%d.%m.%Y").strftime("%Y-%m-%d")
            elif "-" in date_first:
                parsed_date = datetime.strptime(date_first, "%Y-%m-%d").strftime("%Y-%m-%d")
            else:
                parsed_date = date_first
        except:
            parsed_date = date_first

        if parsed_date == today and link not in sent_ads:
            send_telegram_message(
                title=row.get("Title", "Brak tytułu"),
                link=link,
                price=row.get("Price last updated", row.get("Price at first find", "?")),
                image_url=row.get("Image", None)
            )
            sent_ads.add(link)

    sent_ads = sent_ads.intersection(current_ads)
    save_sent_ads(sent_ads)

    # ☁️ Upload sent_ads.json
    upload_to_onedrive(SENT_JSON, token)

if __name__ == "__main__":
    main()
