import requests
import time
import datetime
import os
import pandas as pd
from collections import OrderedDict, defaultdict
from bs4 import BeautifulSoup
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderUnavailable, GeocoderServiceError
from requests.exceptions import ConnectionError, ReadTimeout
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import folium
# -------------------------------
# 🔧 Uncomment the following 2 lines locally to enable loading variables from the .env file
from dotenv import load_dotenv
load_dotenv()

# -------------------------------
# 🔐 Environment variables required in .env:
# ONEDRIVE_CLIENT_ID=your_client_id
# ONEDRIVE_REFRESH_TOKEN=your_refresh_token
# 
# NOTE: The .env file is not pushed to GitHub — add these values to GitHub Secrets as well
# if you want to run the script automatically via GitHub Actions.
# -------------------------------

# -------------------------------
# 🔐 OneDrive authentication variables (stored in .env)
CLIENT_ID = os.environ['ONEDRIVE_CLIENT_ID']
REFRESH_TOKEN = os.environ['ONEDRIVE_REFRESH_TOKEN']
SCOPES = ['offline_access', 'Files.ReadWrite.All']
TOKEN_URL = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'

# -------------------------------
# 📦 App Configuration
BASE_LINK_KRAKOW = 'https://www.otodom.pl/pl/wyniki/sprzedaz/dzialka/malopolskie/krakowski?limit=72&priceMax=250000&areaMin=1300&plotType=%5BBUILDING%2CAGRICULTURAL_BUILDING%5D&by=DEFAULT&direction=DESC'
BASE_LINK_WIELICKI = 'https://www.otodom.pl/pl/wyniki/sprzedaz/dzialka/malopolskie/wielicki?distanceRadius=5&limit=72&priceMax=250000&areaMin=1300&plotType=%5BBUILDING%2CAGRICULTURAL_BUILDING%5D&by=DEFAULT&direction=DESC'
KRAKOW_COORDS = (50.0647, 19.9450)
EXCEL_FOLDER = 'dzialki'
EXCEL_FILENAME = 'otodom_dzialki.xlsx'
MAP_FILE = os.path.join(EXCEL_FOLDER, 'otodom_map_listings.html')
EXCEL_FILE = os.path.join(EXCEL_FOLDER, EXCEL_FILENAME)
SHEET_NAMES = ['powiat krakowski', 'powiat wielicki']
HEADERS = [
    'Title', 'Location', 'Price at first find',
    'Date first found', 'Date last updated',
    'Price last updated', 'Distance from Krakow (km)',
    'Active', 'Link',
    'Latitude', 'Longitude',
]

# Allowed counties around Kraków for better geocoding accuracy
ALLOWED_COUNTIES = ['krakowski', 'wielicki', 'wadowicki', 'chrzanowski', 'olkuski', 'myślenicki']

# Geopy setup
geolocator = Nominatim(user_agent="plot_script")
max_distance_from_Krakow = 50

# -------------------------------
# 📂 Wczytaj miejscowości z pliku TXT
# Format pliku: nazwa_miejscowości,lat,lon (jedna miejscowość w wierszu)
def load_town_coords(filename="town_list.txt"):
    towns = defaultdict(list)
    if os.path.exists(filename):
        with open(filename, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                try:
                    parts = line.split(",")
                    if len(parts) == 3:
                        town = parts[0].strip().lower()
                        lat = float(parts[1].strip())
                        lon = float(parts[2].strip())
                        towns[town].append((lat, lon))
                except Exception as e:
                    print(f"⚠️ Błąd w pliku TXT: {line} -> {e}")
    else:
        print("⚠️ Brak pliku town_list.txt")
    return towns

TOWN_COORDS = load_town_coords()

# -------------------------------
# 🔐 OneDrive token refresh
def authenticate():
    data = {
        'client_id': CLIENT_ID,
        'refresh_token': REFRESH_TOKEN,
        'grant_type': 'refresh_token',
        'scope': 'offline_access Files.ReadWrite.All',
    }
    resp = requests.post(TOKEN_URL, data=data)
    if resp.status_code != 200:
        raise Exception(f"❌ Failed to authenticate: {resp.text}")
    return resp.json()

# -------------------------------
# ☁️ Download file from OneDrive
def download_from_onedrive(file_path, token):
    """Download file from OneDrive and save localy"""
    headers = {
        'Authorization': f"Bearer {token['access_token']}",
    }
    download_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{file_path}:/content'
    r = requests.get(download_url, headers=headers)
    if r.status_code == 200:
        with open(file_path, 'wb') as f:
            f.write(r.content)
        print(f"✅ File downloaded from OneDrive: {file_path}")
    else:
        raise Exception(f"❌ Failed to download file: {r.status_code} {r.text}")

# -------------------------------
# ☁️ Upload file to OneDrive
def upload_to_onedrive(file_path, token):
    headers = {
        'Authorization': f"Bearer {token['access_token']}",
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    with open(file_path, 'rb') as f:
        file_data = f.read()
    upload_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{file_path}:/content'
    r = requests.put(upload_url, headers=headers, data=file_data)
    if r.status_code in (200, 201):
        print(f"✅ File uploaded to OneDrive: {file_path}")
    else:
        print(f"❌ Upload failed: {r.status_code} {r.text}")

# -------------------------------
# 📊 Create Excel file with headers and sheets
def create_excel_with_sheets():
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.remove(wb.active)
        for name in SHEET_NAMES:
            ws = wb.create_sheet(name)
            for i, col in enumerate(HEADERS, 1):
                ws.cell(row=1, column=i, value=col)
                ws.column_dimensions[get_column_letter(i)].width = max(15, len(col) + 2)
        wb.save(EXCEL_FILE)
        print(f"📄 Created Excel file: {EXCEL_FILE}")

# -------------------------------
# 🔍 Utility functions
def parse_price(price_str):
    return int(price_str.replace(' ', '').replace('zł', '').replace('PLN', '').replace(',', '').strip())

def extract_relevant_town(location):
    parts = [part.strip() for part in location.split(',')]
    if parts[0].lower().startswith("ul.") and len(parts) > 1:
        return parts[1]
    else:
        return parts[0]

def safe_geocode(loc: str, max_retries: int = 2, timeout: int = 5):
    for attempt in range(max_retries):
        try:
            return geolocator.geocode(loc, exactly_one=False, timeout=timeout)
        except (GeocoderUnavailable, GeocoderServiceError, ConnectionError, ReadTimeout) as e:
            print(f"⏳ Geocoding failed ({attempt + 1}/{max_retries}): {loc} -> {e}")
            time.sleep(1)
    return None

def get_distance_to_krakow(town, county=""):
    town_key = town.lower()
    if town_key in TOWN_COORDS:  # 🔑 Najpierw sprawdzamy listę TXT
        results = []
        for lat, lon in TOWN_COORDS[town_key]:
            distance = geodesic(KRAKOW_COORDS, (lat, lon)).km
            results.append((round(distance, 2), lat, lon))
        return results  # może być kilka miejscowości o tej samej nazwie

    # jeśli brak w TXT → klasyczna geolokalizacja
    queries = []
    if county and county.lower() in ALLOWED_COUNTIES:
        queries.append(f"{town}, {county} county, Małopolskie, Poland")
    queries.append(f"{town}, Małopolskie, Poland")

    for query in queries:
        places = safe_geocode(query)
        if places:
            results = []
            for p in places:
                if hasattr(p, 'latitude') and hasattr(p, 'longitude'):
                    distance = geodesic(KRAKOW_COORDS, (p.latitude, p.longitude)).km
                    results.append((round(distance, 2), p.latitude, p.longitude))
            if results:
                return results

    print(f"⚠️ Location not found: {town} ({county}) – setting distance as -1 km")
    return [(-1.0, None, None)]

# -------------------------------
# 🧲 Scrape offers from Otodom
HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "pl-PL,pl;q=0.9"
}

def scrape_offers(base_link, name):
    results = []
    today = datetime.date.today().strftime('%Y-%m-%d')
    county = name.replace("powiat ", "")
    try:
        res = requests.get(base_link, headers=HEADERS_HTTP, timeout=30)
        soup = BeautifulSoup(res.text, "html.parser")
        links = list(OrderedDict.fromkeys([
            'https://www.otodom.pl' + a['href'] if a['href'].startswith('/') else a['href']
            for a in soup.select('a[data-cy="listing-item-link"]') if a.get('href')
        ]))
        print(f"🔍 {name}: {len(links)} offers found")

        for idx, url in enumerate(links, 1):
            print(f"➡️ Processing {idx}/{len(links)}: {url}")
            try:
                o = requests.get(url, headers=HEADERS_HTTP, timeout=30)
                s = BeautifulSoup(o.text, "html.parser")
                title = s.find('h1').text.strip() if s.find('h1') else 'No title'
                price = parse_price(s.select_one('strong[data-cy="adPageHeaderPrice"]').text)
                location = s.select_one('div[data-sentry-component="MapLink"] a').text.strip()
                town = extract_relevant_town(location)

                coords_list = get_distance_to_krakow(town, county)  # może być kilka punktów
                for distance, lat, lon in coords_list:
                    results.append({
                        'Title': title,
                        'Location': location,
                        'Price at first find': price,
                        'Date first found': today,
                        'Date last updated': today,
                        'Price last updated': price,
                        'Distance from Krakow (km)': distance,
                        'Active': True,
                        'Link': url,
                        'Latitude': lat,
                        'Longitude': lon
                    })
                time.sleep(2)
            except Exception as e:
                print(f"❌ Skipping offer due to error: {e}")
    except Exception as e:
        print(f"❌ Scraping error: {e}")
    return results

# -------------------------------
# 🧾 Update Excel sheet with offers
def update_sheet(results, sheet_name):
    """Update Excel sheet with new and existing offers"""
    today = datetime.date.today().strftime('%Y-%m-%d')
    df_new = pd.DataFrame(results).dropna(how='all')

    if os.path.exists(EXCEL_FILE):
        xls = pd.ExcelFile(EXCEL_FILE)
        df_old = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name) if sheet_name in xls.sheet_names else pd.DataFrame(columns=HEADERS)
    else:
        df_old = pd.DataFrame(columns=HEADERS)

    for r in results:
        match = df_old[(df_old['Title'] == r['Title']) & (df_old['Price last updated'] == r['Price last updated'])]
        if not match.empty:
            i = match.index[0]
            df_old.at[i, 'Date last updated'] = today
            df_old.at[i, 'Active'] = True
        else:
            df_old = pd.concat([df_old, pd.DataFrame([r])], ignore_index=True)

    existing_links = [r['Link'] for r in results]
    if 'Link' in df_old.columns:
        df_old['Active'] = df_old['Link'].apply(lambda x: x in existing_links)

    wb = load_workbook(EXCEL_FILE)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    wb.save(EXCEL_FILE)
    wb.close()

    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_old.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        for i, col in enumerate(df_old.columns, 1):
            max_length = max(
                df_old[col].astype(str).map(len).max(),
                len(col)
            )
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    print(f"✅ Saved {len(df_old)} offers to sheet '{sheet_name}'")

# -------------------------------
# 🗺️ Generate map from Excel data
def generate_map(df):
    """Generate interactive map with markers grouped by coordinates"""
    m = folium.Map(location=KRAKOW_COORDS, zoom_start=10)

    # Kraków reference marker
    folium.Marker(
        location=KRAKOW_COORDS,
        popup=folium.Popup("<b>Kraków</b><br>Reference point", max_width=200),
        tooltip="Kraków",
        icon=folium.Icon(color="purple", icon="star", prefix="fa")
    ).add_to(m)

    # Group offers by coordinates
    marker_groups = defaultdict(list)
    for _, row in df.iterrows():
        if not row.get("Active", True):
            continue
        lat, lon = row.get("Latitude"), row.get("Longitude")
        if pd.isna(lat) or pd.isna(lon):
            continue
        marker_groups[(lat, lon)].append(row)

    # Add grouped markers
    for (lat, lon), listings in marker_groups.items():
        if len(listings) == 1:
            row = listings[0]
            popup_html = f"""
            <b>{row['Title']}</b><br>
            {row['Location']}<br>
            {row.get('Price last updated', '')} PLN<br>
            <a href='{row['Link']}' target='_blank'>View listing</a>
            """
            tooltip = row['Title']
        else:
            popup_html = f"<b>{len(listings)} listings</b><br><ul>"
            for r in listings:
                popup_html += f"<li><a href='{r['Link']}' target='_blank'>{r['Title']}</a> – {r['Price last updated']} PLN</li>"
            popup_html += "</ul>"
            tooltip = f"{len(listings)} listings in this location"

        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_html, max_width=300),
            tooltip=tooltip,
            icon=folium.Icon(color="green", icon="home", prefix="fa")
        ).add_to(m)

    m.save(MAP_FILE)
    print(f"🗺️ Map saved to: {MAP_FILE}")

# -------------------------------
# 🚀 MAIN Function
def main():
    if CLIENT_ID and REFRESH_TOKEN:
        token = authenticate()
        # Download newest version from OneDrive
        download_from_onedrive(EXCEL_FILE, token)
    else:
        print("⚠️ OneDrive credentials not found. Using local Excel copy.")

    create_excel_with_sheets()
    update_sheet(scrape_offers(BASE_LINK_KRAKOW, 'powiat krakowski'), 'powiat krakowski')
    update_sheet(scrape_offers(BASE_LINK_WIELICKI, 'powiat wielicki'), 'powiat wielicki')

    df_combined = pd.concat([
        pd.read_excel(EXCEL_FILE, sheet_name='powiat krakowski'),
        pd.read_excel(EXCEL_FILE, sheet_name='powiat wielicki')
    ], ignore_index=True)

    generate_map(df_combined)

    # Upload updated file on OneDrive
    if CLIENT_ID and REFRESH_TOKEN:
        print(f"📦 Uploading updated Excel and map to OneDrive...")
        upload_to_onedrive(EXCEL_FILE, token)
        upload_to_onedrive(MAP_FILE, token)
    else:
        print("⚠️ OneDrive credentials not found. Skipping upload.")


# -------------------------------
# 🚀 MAIN SCRIPT ENTRY POINT
if __name__ == "__main__":
    main()
