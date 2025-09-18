import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import openpyxl
from openpyxl.utils import get_column_letter
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
import folium
from collections import defaultdict
from datetime import date

# -------------------------------
# Uncomment locally to enable loading variables from the .env file
from dotenv import load_dotenv
load_dotenv()

# -------------------------------
# Environment variables required in .env:
# ONEDRIVE_CLIENT_ID=your_client_id
# ONEDRIVE_REFRESH_TOKEN=your_refresh_token
#
# NOTE: The .env file is not pushed to GitHub — add these values to GitHub Secrets as well
# if you want to run the script automatically via GitHub Actions.
# -------------------------------

# Constants
KRAKOW_COORDS = (50.0647, 19.9450)
MAX_DISTANCE_KM = 50
MAX_PAGES = 20
EXCEL_FOLDER = 'dzialki'
EXCEL_FILENAME = 'nieruchomosci_online_dzialki.xlsx'
EXCEL_FILE = os.path.join(EXCEL_FOLDER, EXCEL_FILENAME)
MAP_FILE = os.path.join(EXCEL_FOLDER, 'nieruchomosci_online_map_listings.html')
HEADERS = {"User-Agent": "Mozilla/5.0"}
BASE_URL = "https://www.nieruchomosci-online.pl/szukaj.html?3,dzialka,sprzedaz,,Krak%C3%B3w:5600,,,25,-250000,1150,,,,,,,,,,,,,1"

# OneDrive credentials
CLIENT_ID = os.environ.get('ONEDRIVE_CLIENT_ID')
REFRESH_TOKEN = os.environ.get('ONEDRIVE_REFRESH_TOKEN')
TOKEN_URL = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'

# Allowed counties for Małopolska
ALLOWED_COUNTIES = ['krakowski', 'wielicki', 'wadowicki', 'chrzanowski', 'olkuski', 'myślenicki']
COUNTY_COORDS = {
    'krakowski': (50.0647, 19.9450),
    'wielicki': (49.9871, 20.0644),
    'wadowicki': (49.8833, 19.4881),
    'chrzanowski': (50.1376, 19.3988),
    'olkuski': (50.2810, 19.5653),
    'myślenicki': (49.8333, 19.9333)
}

results = []
geolocator = Nominatim(user_agent="dzialki_locator")
today = date.today().isoformat()

# -------------------------------
# Load towns list from file
def load_towns(file_path="town_list.txt"):
    """Load town coordinates from file into a dictionary of lists"""
    towns = defaultdict(list)
    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split("|")
                if len(parts) == 3:
                    town, lat, lon = parts
                    towns[town.lower()].append((float(lat), float(lon)))
    print(f"Loaded {sum(len(v) for v in towns.values())} town coordinates from {file_path}")
    return towns

TOWN_COORDS = load_towns("town_list.txt")

# -------------------------------
# OneDrive authentication
def authenticate():
    """Authenticate with OneDrive API using refresh token"""
    data = {
        'client_id': CLIENT_ID,
        'refresh_token': REFRESH_TOKEN,
        'grant_type': 'refresh_token',
        'scope': 'offline_access Files.ReadWrite.All',
    }
    resp = requests.post(TOKEN_URL, data=data)
    if resp.status_code != 200:
        raise Exception(f"❌ Failed to authenticate: {resp.text}")
    return resp.json()

# -------------------------------
# Upload file to OneDrive
def upload_to_onedrive(file_path, token):
    """Upload a file to OneDrive root directory"""
    headers = {
        'Authorization': f"Bearer {token['access_token']}",
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    with open(file_path, 'rb') as f:
        file_data = f.read()
    upload_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{file_path}:/content'
    r = requests.put(upload_url, headers=headers, data=file_data)
    if r.status_code in (200, 201):
        print(f"✅ File uploaded to OneDrive: {file_path}")
    else:
        print(f"❌ Upload failed: {r.status_code} {r.text}")

# -------------------------------
# Download file from OneDrive
def download_from_onedrive(file_path, token):
    """Download a file from OneDrive and save it locally"""
    headers = {
        'Authorization': f"Bearer {token['access_token']}",
    }
    download_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{file_path}:/content'
    r = requests.get(download_url, headers=headers)
    if r.status_code == 200:
        with open(file_path, 'wb') as f:
            f.write(r.content)
        print(f"✅ File downloaded from OneDrive: {file_path}")
    else:
        print(f"⚠️ Failed to download file from OneDrive: {r.status_code} {r.text}")


# -------------------------------
# Get distance from Kraków with local list check, county filter and geopy fallback
def get_distance_from_krakow(location, max_retries=3):
    """Return list of (distance, lat, lon) if within allowed counties and distance"""
    town = location.split("(")[0].strip().lower()
    results = []

    # First check in town_list.txt
    if town in TOWN_COORDS:
        for lat, lon in TOWN_COORDS[town]:
            distance = round(geodesic(KRAKOW_COORDS, (lat, lon)).km, 2)
            if distance <= MAX_DISTANCE_KM:
                results.append((distance, lat, lon))
            else:
                print(f"⚠️ {town} (from file) found but too far: {distance} km")
        if results:
            return results

    # Check if it's a county
    for county in ALLOWED_COUNTIES:
        if county in town:
            lat, lon = COUNTY_COORDS[county]
            distance = round(geodesic(KRAKOW_COORDS, (lat, lon)).km, 2)
            if distance <= MAX_DISTANCE_KM:
                return [(distance, lat, lon)]
            else:
                print(f"⚠️ {town} (county) found but too far: {distance} km")
                return []

    # Fallback to geopy
    for attempt in range(max_retries):
        try:
            geo = geolocator.geocode(f"{town}, Lesser Poland Voivodeship, Poland", timeout=10)
            if geo:
                coords = (geo.latitude, geo.longitude)
                distance = round(geodesic(KRAKOW_COORDS, coords).km, 2)
                if distance <= MAX_DISTANCE_KM:
                    return [(distance, coords[0], coords[1])]
                else:
                    print(f"⚠️ {town} (geopy) found but too far: {distance} km")
                    return []
        except Exception as e:
            print(f"⚠️ Geocoding attempt {attempt+1} failed for {town}: {e}")
    return []

# -------------------------------
# Main scraping function
def main():
    os.makedirs(EXCEL_FOLDER, exist_ok=True)
    total_raw = total_unique = total_geocoded = 0

    # -------------------------------
    # Load existing Excel to keep history
    if CLIENT_ID and REFRESH_TOKEN:
        print("☁️ OneDrive credentials found. Downloading latest Excel copy...")
        token = authenticate()
        download_from_onedrive(EXCEL_FILE, token)
    else:
        print("⚠️ OneDrive credentials not found. Using local Excel copy.")

    if os.path.exists(EXCEL_FILE):
        df_existing = pd.read_excel(EXCEL_FILE)
        print(f"✅ Loaded existing Excel with {len(df_existing)} rows")
    else:
        df_existing = pd.DataFrame(columns=[
            'Title','Location','Price at first find','Date first found','Date last updated',
            'Price last updated','Distance from Krakow (km)','Active','Link','Latitude','Longitude'
        ])
        print("ℹ️ No existing Excel file found, creating new DataFrame")

    # Load existing Excel to keep history
    if os.path.exists(EXCEL_FILE):
        df_existing = pd.read_excel(EXCEL_FILE)
        print(f"✅ Loaded existing Excel with {len(df_existing)} rows")
    else:
        df_existing = pd.DataFrame(columns=[
            'Title','Location','Price at first find','Date first found','Date last updated',
            'Price last updated','Distance from Krakow (km)','Active','Link','Latitude','Longitude'
        ])
        print("ℹ️ No existing Excel file found, creating new DataFrame")

    current_links = set()  # track currently found links

    for page in range(1, MAX_PAGES + 1):
        url = BASE_URL if page == 1 else f"{BASE_URL}&p={page}"
        print(f"\n🌐 Fetching page {page}: {url}")
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            print(f"❌ Failed to load page {page}")
            break

        soup = BeautifulSoup(response.text, "html.parser")
        raw_links = soup.select("h2.name a")
        total_raw += len(raw_links)
        print(f"✅ Found {len(raw_links)} raw listings on page {page}")

        seen_links = set()
        clean_listings = []
        for a in raw_links:
            link = a.get("href")
            if not link or link in seen_links:
                continue
            seen_links.add(link)
            parent = a.find_parent("div", class_=lambda x: x and ("tile-inner" in x or "tertiary" in x))
            if parent:
                clean_listings.append(parent)
        total_unique += len(clean_listings)
        print(f"🔎 After deduplication: {len(clean_listings)} unique listings on page {page}")

        for listing in clean_listings:
            try:
                title_tag = listing.find("h2", class_="name")
                title = title_tag.text.strip() if title_tag else "No title"
                link = title_tag.find("a")["href"] if title_tag and title_tag.find("a") else "No link"
                current_links.add(link)

                price_tag = listing.find("p", class_="title-a")
                price_span = price_tag.find("span") if price_tag else None
                price = price_span.text.strip() if price_span else "No price"

                location_p = listing.find("p", class_="province")
                location = " ".join([el.text.strip() for el in location_p.find_all(['a', 'span'])]) if location_p else "No location"

                coords_list = get_distance_from_krakow(location)
                if coords_list:
                    for distance, lat, lon in coords_list:
                        total_geocoded += 1

                        existing_row = df_existing[df_existing['Link'] == link]
                        if not existing_row.empty:
                            first_date = existing_row.iloc[0]['Date first found']
                            first_price = existing_row.iloc[0]['Price at first find']
                        else:
                            first_date = today
                            first_price = price

                        results.append({
                            'Title': title,
                            'Location': location,
                            'Price at first find': first_price,
                            'Date first found': first_date,
                            'Date last updated': today,
                            'Price last updated': price,
                            'Distance from Krakow (km)': distance,
                            'Active': True,
                            'Link': link,
                            'Latitude': lat,
                            'Longitude': lon
                        })
            except Exception as e:
                print(f"❗ Error parsing listing: {e}")
                continue
        time.sleep(1)

    # Mark previously existing listings not found in current scrape as inactive
    df_existing_links = set(df_existing['Link'])
    missing_links = df_existing_links - current_links
    print(f"ℹ️ {len(missing_links)} listings not found in this scrape will be marked inactive")
    for link in missing_links:
        row = df_existing['Link'] == link
        df_existing.loc[row, 'Active'] = False

    # Combine new results with existing inactive ones
    df_new = pd.DataFrame(results)
    df_combined = pd.concat([df_existing[df_existing['Active'] == False], df_new], ignore_index=True)
    df_combined = df_combined.drop_duplicates(subset=['Link','Latitude','Longitude'])

    # Save to Excel
    df_combined.to_excel(EXCEL_FILE, index=False)

    # Auto-fit Excel columns
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    wb.save(EXCEL_FILE)

    # Create map
    m = folium.Map(location=KRAKOW_COORDS, zoom_start=10)
    folium.Marker(location=KRAKOW_COORDS, popup="Kraków - Reference Point", tooltip="Kraków",
                  icon=folium.Icon(color="purple")).add_to(m)

    marker_groups = defaultdict(list)
    for _, row in df_combined.iterrows():
        coord = (row["Latitude"], row["Longitude"])
        marker_groups[coord].append(row)

    for coord, listings in marker_groups.items():
        if len(listings) == 1:
            l = listings[0]
            popup_html = f"<b>{l['Title']}</b><br>{l['Location']}<br>{l['Price last updated']}<br><a href='{l['Link']}' target='_blank'>View Listing</a>"
            tooltip = l["Title"]
        else:
            popup_html = f"<b>{len(listings)} listings</b><br><ul>"
            for l in listings:
                popup_html += f"<li><a href='{l['Link']}' target='_blank'>{l['Title']}</a> – {l['Price last updated']}</li>"
            popup_html += "</ul>"
            tooltip = f"{len(listings)} listings at same location"

        folium.Marker(location=coord, popup=folium.Popup(popup_html, max_width=300),
                      tooltip=tooltip, icon=folium.Icon(color="orange", icon="home")).add_to(m)

    m.save(MAP_FILE)
    print(f"\n✅ Data saved to: {EXCEL_FILE}")
    print(f"🗺️ Map saved to: {MAP_FILE}")

    if CLIENT_ID and REFRESH_TOKEN:
        print(f"📦 Done. Uploading {EXCEL_FILE} and map to OneDrive...")
        token = authenticate()
        upload_to_onedrive(EXCEL_FILE, token)
        upload_to_onedrive(MAP_FILE, token)
    else:
        print("⚠️ OneDrive credentials not found. Skipping upload.")

    print(f"📊 SUMMARY:\n   Raw listings found: {total_raw}\n   Unique listings: {total_unique}\n   Geocoded listings: {total_geocoded}")

if __name__ == "__main__":
    main()
